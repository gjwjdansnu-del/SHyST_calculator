#!/usr/bin/env python3
"""
SHyST 실험 체크리스트 엑셀 데이터를 JSON으로 변환
"""

import openpyxl
import json
from datetime import datetime

def parse_date(date_str):
    """날짜 문자열을 YYYY-MM-DD 형식으로 변환"""
    if not date_str:
        return ''
    
    date_str = str(date_str).strip()
    
    # YYMMDD 형식 (예: 240826 -> 2024-08-26)
    if len(date_str) == 6 and date_str.isdigit():
        year = '20' + date_str[:2]
        month = date_str[2:4]
        day = date_str[4:6]
        return f"{year}-{month}-{day}"
    
    return date_str

def safe_float(value):
    """안전하게 float로 변환"""
    if value is None or value == '':
        return None
    try:
        return float(value)
    except:
        return None

def safe_int(value):
    """안전하게 int로 변환"""
    if value is None or value == '':
        return None
    try:
        return int(value)
    except:
        return None

def safe_str(value):
    """안전하게 str로 변환"""
    if value is None:
        return ''
    return str(value).strip()

def parse_experiment_row(ws, row_num):
    """엑셀 행에서 실험 데이터 파싱"""
    
    # 실험 번호 확인
    exp_num = ws.cell(row_num, 1).value
    if exp_num is None or exp_num == '#':
        return None
    
    exp_num_str = str(exp_num).strip()
    if exp_num_str.startswith('0 '):  # 예시 행 제외
        return None
    
    try:
        exp_number = int(exp_num_str)
    except:
        return None
    
    # 데이터 파싱
    experiment = {
        'id': None,
        'expNumber': exp_number,
        'status': 'completed',
        'createdAt': datetime.now().isoformat(),
        'updatedAt': datetime.now().isoformat(),
        
        'before': {
            'expInfo': {
                'name': safe_str(ws.cell(row_num, 2).value),
                'date': parse_date(ws.cell(row_num, 3).value),
                'testModel': safe_str(ws.cell(row_num, 4).value),
                'objective': safe_str(ws.cell(row_num, 5).value),
                'targetMach': safe_float(ws.cell(row_num, 6).value)
            },
            'shystSetting': {
                'airPressure': safe_float(ws.cell(row_num, 7).value),
                'airTemp': safe_float(ws.cell(row_num, 8).value),
                'airHumidity': safe_float(ws.cell(row_num, 9).value),
                'driverGas': safe_str(ws.cell(row_num, 10).value),
                'boosterPressure': safe_float(ws.cell(row_num, 11).value),
                'firstDiaphragm': safe_str(ws.cell(row_num, 12).value),
                'secondDiaphragm': safe_str(ws.cell(row_num, 13).value),
                'drivenGas': safe_str(ws.cell(row_num, 14).value),
                'drivenPressure': safe_float(ws.cell(row_num, 15).value),
                'drivenTemp': safe_float(ws.cell(row_num, 16).value),
                'vacuumGauge': safe_float(ws.cell(row_num, 17).value),
                'daqSampling': safe_float(ws.cell(row_num, 18).value)
            },
            'visualizationSetting': {
                'method': safe_str(ws.cell(row_num, 19).value),
                'target': safe_str(ws.cell(row_num, 20).value)
            },
            'cameraSetting': {
                'model': safe_str(ws.cell(row_num, 21).value),
                'fps': safe_float(ws.cell(row_num, 22).value),
                'width': safe_int(ws.cell(row_num, 23).value),
                'height': safe_int(ws.cell(row_num, 24).value),
                'lensFocal': safe_str(ws.cell(row_num, 25).value),
                'exposeTime': safe_float(ws.cell(row_num, 26).value)
            }
        },
        
        'after': {
            'labviewLog': {
                'p1_avg': safe_float(ws.cell(row_num, 27).value),
                't1_avg': safe_float(ws.cell(row_num, 28).value),
                'p4_avg': safe_float(ws.cell(row_num, 29).value),
                'p4_std': safe_float(ws.cell(row_num, 30).value),
                't4_avg': safe_float(ws.cell(row_num, 31).value),
                'p5_avg': safe_float(ws.cell(row_num, 32).value),
                'p5_std': safe_float(ws.cell(row_num, 33).value),
                'testTime': safe_float(ws.cell(row_num, 34).value),
                'shockSpeed': safe_float(ws.cell(row_num, 35).value),
                'outputDelayTime': safe_float(ws.cell(row_num, 36).value),
                'outputReadyTime': safe_float(ws.cell(row_num, 37).value)
            },
            'rawDataFiles': [],
            'sensorCalibrations': [],
            'selectedTestTime': {
                'start': None,
                'end': None,
                'duration': safe_float(ws.cell(row_num, 34).value)
            }
        },
        
        'calculation': {
            'method': 'estcn',
            'stages': {
                'stage1': parse_stage(ws, row_num, 38) if ws.cell(row_num, 38).value else None,
                'stage2': parse_stage(ws, row_num, 50) if ws.cell(row_num, 50).value else None,
                'stage5': parse_stage(ws, row_num, 62) if ws.cell(row_num, 62).value else None,
                'stage5s': parse_stage(ws, row_num, 74) if ws.cell(row_num, 74).value else None,
                'stage6': parse_stage(ws, row_num, 86) if ws.cell(row_num, 86).value else None,
                'stage7': parse_stage_7(ws, row_num, 98) if ws.cell(row_num, 98).value else None
            }
        }
    }
    
    return experiment

def parse_stage(ws, row_num, col_start):
    """Stage 데이터 파싱 (12개 컬럼)"""
    internal_energy = safe_float(ws.cell(row_num, col_start + 3).value)
    velocity = safe_float(ws.cell(row_num, col_start + 10).value)
    return {
        'p': safe_float(ws.cell(row_num, col_start).value),
        't': safe_float(ws.cell(row_num, col_start + 1).value),
        'rho': safe_float(ws.cell(row_num, col_start + 2).value),
        # The fourth legacy-Excel column is thermodynamic internal energy,
        # not flow velocity. Logger JSON uses `u` as a velocity alias.
        'e': internal_energy,
        'u': velocity,
        'h': safe_float(ws.cell(row_num, col_start + 4).value),
        'R': safe_float(ws.cell(row_num, col_start + 5).value),
        'gamma': safe_float(ws.cell(row_num, col_start + 6).value),
        'cp': safe_float(ws.cell(row_num, col_start + 7).value),
        'a': safe_float(ws.cell(row_num, col_start + 8).value),
        's': safe_float(ws.cell(row_num, col_start + 9).value),
        'V': velocity,
        'M': safe_float(ws.cell(row_num, col_start + 11).value)
    }

def parse_stage_7(ws, row_num, col_start):
    """Stage 7 데이터 파싱 (추가 정보 포함)"""
    stage7 = parse_stage(ws, row_num, col_start)
    
    # Re7, h_tot7 추가 (Col 110, 111)
    # Legacy workbook stores raw Re/m and J/kg, not values scaled by 1e6.
    stage7['Re_unit'] = safe_float(ws.cell(row_num, 110).value)
    stage7['H0'] = safe_float(ws.cell(row_num, 111).value)
    stage7['Re_unit_e6'] = (
        stage7['Re_unit'] / 1e6 if stage7['Re_unit'] is not None else None
    )
    stage7['H0_MJ'] = (
        stage7['H0'] / 1e6 if stage7['H0'] is not None else None
    )
    
    # 여기까지만 저장 (Col 111 h_tot7까지)
    # Col 113 이후 (pcc #, image format, 천이레이놀즈수 등)는 제외
    
    return stage7

def main():
    print("🚀 SHyST 실험 데이터 변환 시작...")
    
    # 엑셀 파일 로드
    wb = openpyxl.load_workbook('../SHyST Exp Check List ver1.xlsx', data_only=True)
    ws = wb.worksheets[0]
    
    print(f"📊 엑셀 파일 로드 완료: {ws.max_row} 행")
    
    # 데이터 파싱
    experiments = []
    
    # Row 4부터 시작 (Row 1-3은 헤더)
    for row_num in range(4, ws.max_row + 1):
        exp = parse_experiment_row(ws, row_num)
        if exp:
            experiments.append(exp)
            print(f"✅ 실험 #{exp['expNumber']} 파싱 완료 - {exp['before']['expInfo']['name']} ({exp['before']['expInfo']['date']})")
    
    print(f"\n📦 총 {len(experiments)}개 실험 데이터 파싱 완료")
    
    # JSON 파일로 저장
    output_file = 'experiments_data.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(experiments, f, ensure_ascii=False, indent=2)
    
    print(f"💾 JSON 파일 저장 완료: {output_file}")
    print(f"\n🎉 변환 완료! 웹 페이지에서 불러올 수 있습니다.")

if __name__ == '__main__':
    main()
